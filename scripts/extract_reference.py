import json
import sys
from pathlib import Path

import openpyxl


def text(value):
    return "" if value is None else str(value).strip()


def number(value, default=0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


source = Path(sys.argv[1] if len(sys.argv) > 1 else "/Users/yanxingyu/ai/跨境电商ERP_WPS完整业务版.xlsx")
target = Path(sys.argv[2] if len(sys.argv) > 2 else "packages/core/src/reference-data.json")
workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)

sku_sheet = workbook["01_SKU映射"]
skus = []
stocks = []
for row in range(5, sku_sheet.max_row + 1):
    code = text(sku_sheet.cell(row, 1).value)
    if not code:
        continue
    skus.append(
        {
            "id": f"sku-{len(skus) + 1}",
            "code": code,
            "sellerSku": text(sku_sheet.cell(row, 3).value),
            "platformSku": text(sku_sheet.cell(row, 4).value),
            "platformSkc": text(sku_sheet.cell(row, 5).value),
            "platformSpu": text(sku_sheet.cell(row, 6).value),
            "name": text(sku_sheet.cell(row, 7).value),
            "spec": text(sku_sheet.cell(row, 8).value),
            "sellerCode": text(sku_sheet.cell(row, 9).value),
            "shippingName": text(sku_sheet.cell(row, 10).value),
            "shippingMethod": text(sku_sheet.cell(row, 11).value),
            "imageUrl": text(sku_sheet.cell(row, 12).value),
            "supplier": text(sku_sheet.cell(row, 13).value),
            "purchaseLink": text(sku_sheet.cell(row, 14).value),
            "purchasePrice": number(sku_sheet.cell(row, 15).value),
            "leadTimeDays": int(number(sku_sheet.cell(row, 16).value, 15)),
            "safetyDays": 7,
            "safetyStock": int(number(sku_sheet.cell(row, 17).value)),
            "reorderPoint": int(number(sku_sheet.cell(row, 18).value)),
            "targetStock": int(number(sku_sheet.cell(row, 19).value)),
            "confirmStatus": text(sku_sheet.cell(row, 21).value),
            "owner": text(sku_sheet.cell(row, 22).value),
        }
    )
    initial = int(number(sku_sheet.cell(row, 20).value))
    stocks.append({"warehouse": "义乌一仓", "skuCode": code, "quantity": initial})

order_sheet = workbook["02_SHEIN订单粘贴"]
sku_map = {item["sellerSku"]: item["code"] for item in skus}
orders = []
for row in range(5, min(order_sheet.max_row, 204) + 1):
    order_no = text(order_sheet.cell(row, 1).value)
    if not order_no:
        continue
    seller_sku = text(order_sheet.cell(row, 9).value)
    sku_code = sku_map.get(seller_sku)
    orders.append(
        {
            "id": f"ord-{len(orders) + 1}",
            "orderNo": order_no,
            "createdAt": text(order_sheet.cell(row, 2).value),
            "shipBy": text(order_sheet.cell(row, 4).value),
            "sellerSku": seller_sku,
            "skuCode": sku_code,
            "productName": text(order_sheet.cell(row, 6).value),
            "spec": text(order_sheet.cell(row, 7).value),
            "quantity": 1,
            "price": number(order_sheet.cell(row, 13).value),
            "currency": text(order_sheet.cell(row, 15).value),
            "country": text(order_sheet.cell(row, 17).value),
            "warehouse": "义乌一仓",
            "status": "待发货" if sku_code else "异常",
        }
    )

target.parent.mkdir(parents=True, exist_ok=True)
target.write_text(
    json.dumps(
        {"skus": skus, "orders": orders, "stocks": stocks, "shipmentDrafts": []},
        ensure_ascii=False,
        indent=2,
    ),
    encoding="utf-8",
)
print(f"extracted {len(skus)} SKUs and {len(orders)} orders to {target}")
